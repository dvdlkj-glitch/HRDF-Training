"""
One-Click Monthly Report Generator  |  一鍵月報產生器
=====================================================
Excel x AI Operational Excellence — Automation Lab demo

WHAT IT DOES (做什麼):
  Reads a raw transactions file (sample_sales_data.xlsx) and, in one run,
  produces a polished monthly report (Monthly_Report.xlsx) with:
    1. A KPI summary band            (總覽 KPI)
    2. Revenue by Category pivot     (各類別營收樞紐)
    3. Revenue by Region pivot       (各區域營收)
    4. Top 5 products                (前五名產品)
    5. A bar chart                   (長條圖)

WHY IT MATTERS (價值):
  The same monthly task that used to take hours of copy-paste and manual
  formatting now finishes in seconds — every month, identical and error-free.

HOW IT WAS BUILT (如何產生):
  You describe the report you want to Claude in plain language; Claude writes
  this Python. You only press "Run". No coding background required.

Run:  python report_generator.py
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

HERE = Path(__file__).parent
RAW_FILE = HERE / "sample_sales_data.xlsx"
OUT_FILE = HERE / "Monthly_Report.xlsx"

# ---- Brand colours (match the course / website) ----
FOREST = "0A2F2A"
EMERALD = "1F9D57"
LIME = "8DC63F"
CREAM = "F4F7F0"


def ensure_sample_data():
    """若無範例資料，自動產生一份，讓示範一定跑得起來."""
    if RAW_FILE.exists():
        return
    import random
    random.seed(42)
    regions = ["North", "South", "Central", "East"]
    cats = ["Hardware", "Software", "Services", "Accessories"]
    products = {
        "Hardware": ["Laptop", "Monitor", "Printer"],
        "Software": ["CRM Licence", "Antivirus", "Office Suite"],
        "Services": ["Support Plan", "Installation", "Training"],
        "Accessories": ["Mouse", "Keyboard", "Cable"],
    }
    people = ["Aishah", "David", "Kumar", "Mei Ling", "Faizal"]
    rows = []
    for day in range(1, 29):
        for _ in range(random.randint(6, 10)):
            cat = random.choice(cats)
            prod = random.choice(products[cat])
            units = random.randint(1, 25)
            unit_price = {"Hardware": 1500, "Software": 400,
                          "Services": 800, "Accessories": 60}[cat]
            unit_price = int(unit_price * random.uniform(0.8, 1.2))
            rows.append({
                "Date": pd.Timestamp(2026, 5, day),
                "Region": random.choice(regions),
                "Category": cat,
                "Product": prod,
                "Salesperson": random.choice(people),
                "Units": units,
                "Revenue": units * unit_price,
            })
    pd.DataFrame(rows).to_excel(RAW_FILE, index=False)
    print(f"  + Created sample data: {RAW_FILE.name}")


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=EMERALD)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws):
    for col in ws.columns:
        cells = [c for c in col if not isinstance(c, type(None))]
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width + 4


def build_report():
    ensure_sample_data()
    df = pd.read_excel(RAW_FILE)
    month_label = pd.to_datetime(df["Date"]).dt.strftime("%B %Y").mode()[0]

    # ---- Aggregations (pandas does the heavy lifting) ----
    by_cat = (df.groupby("Category")["Revenue"].sum()
                .sort_values(ascending=False).reset_index())
    by_region = (df.groupby("Region")["Revenue"].sum()
                   .sort_values(ascending=False).reset_index())
    top_products = (df.groupby("Product")
                      .agg(Units=("Units", "sum"), Revenue=("Revenue", "sum"))
                      .sort_values("Revenue", ascending=False).head(5).reset_index())

    total_rev = int(df["Revenue"].sum())
    total_units = int(df["Units"].sum())
    orders = len(df)
    avg_order = int(total_rev / orders)

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"
    ws.sheet_view.showGridLines = False
    # Fit all columns onto one page when printing / exporting to PDF
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    # ---- Title band ----
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"Monthly Sales Report  |  月度銷售報表  —  {month_label}"
    t.font = Font(bold=True, size=16, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=FOREST)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ---- KPI band ----
    kpis = [("Total Revenue 總營收", f"RM {total_rev:,}"),
            ("Total Units 總數量", f"{total_units:,}"),
            ("Orders 訂單數", f"{orders:,}"),
            ("Avg Order 平均單筆", f"RM {avg_order:,}")]
    col = 1
    for label, val in kpis:
        c = ws.cell(row=3, column=col, value=label)
        c.font = Font(size=9, color="5F7A6C")
        v = ws.cell(row=4, column=col, value=val)
        v.font = Font(bold=True, size=14, color=FOREST)
        for r in (3, 4):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=CREAM)
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
        col += 1

    # ---- Revenue by Category table ----
    start = 6
    ws.cell(row=start, column=1, value="Revenue by Category 各類別營收").font = Font(bold=True, size=12, color=EMERALD)
    hr = start + 1
    ws.cell(row=hr, column=1, value="Category"); ws.cell(row=hr, column=2, value="Revenue (RM)")
    style_header(ws.cell(row=hr, column=1)); style_header(ws.cell(row=hr, column=2))
    for i, row in by_cat.iterrows():
        ws.cell(row=hr + 1 + i, column=1, value=row["Category"])
        ws.cell(row=hr + 1 + i, column=2, value=int(row["Revenue"])).number_format = "#,##0"

    # ---- Revenue by Region ----
    rstart = hr + len(by_cat) + 3
    ws.cell(row=rstart, column=1, value="Revenue by Region 各區域營收").font = Font(bold=True, size=12, color=EMERALD)
    rh = rstart + 1
    ws.cell(row=rh, column=1, value="Region"); ws.cell(row=rh, column=2, value="Revenue (RM)")
    style_header(ws.cell(row=rh, column=1)); style_header(ws.cell(row=rh, column=2))
    for i, row in by_region.iterrows():
        ws.cell(row=rh + 1 + i, column=1, value=row["Region"])
        ws.cell(row=rh + 1 + i, column=2, value=int(row["Revenue"])).number_format = "#,##0"

    # ---- Top 5 products ----
    pstart = rh + len(by_region) + 3
    ws.cell(row=pstart, column=1, value="Top 5 Products 前五名產品").font = Font(bold=True, size=12, color=EMERALD)
    ph = pstart + 1
    for j, head in enumerate(["Product", "Units", "Revenue (RM)"], start=1):
        ws.cell(row=ph, column=j, value=head); style_header(ws.cell(row=ph, column=j))
    for i, row in top_products.iterrows():
        ws.cell(row=ph + 1 + i, column=1, value=row["Product"])
        ws.cell(row=ph + 1 + i, column=2, value=int(row["Units"]))
        ws.cell(row=ph + 1 + i, column=3, value=int(row["Revenue"])).number_format = "#,##0"

    # ---- thin borders on data area ----
    thin = Side(style="thin", color="D7E6D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=hr, max_row=ph + len(top_products), min_col=1, max_col=3):
        for c in row:
            if c.value is not None:
                c.border = border

    # ---- Bar chart (placed below the tables so it always prints) ----
    chart = BarChart()
    chart.title = "Revenue by Category 各類別營收"
    chart.type = "col"
    chart.legend = None
    data = Reference(ws, min_col=2, min_row=hr, max_row=hr + len(by_cat))
    cats_ref = Reference(ws, min_col=1, min_row=hr + 1, max_row=hr + len(by_cat))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height, chart.width = 8, 15
    ws.add_chart(chart, f"A{ph + len(top_products) + 3}")

    autosize(ws)
    wb.save(OUT_FILE)
    print(f"  + Report saved: {OUT_FILE.name}")
    print(f"    Month: {month_label} | Revenue: RM {total_rev:,} | Orders: {orders}")


if __name__ == "__main__":
    print("One-Click Monthly Report Generator 一鍵月報產生器")
    build_report()
    print("Done. Open Monthly_Report.xlsx to view the result.")
